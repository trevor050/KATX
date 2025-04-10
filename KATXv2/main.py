import openpyxl
import os
import re
import io
from PIL import Image # Pillow is imported as PIL

# --- Configuration ---
EXCEL_FILE_PATH = 'Value List Editor.xlsx' # Path relative to project root
OUTPUT_IMAGE_DIR = 'public/images'        # Output directory for images (relative to root)
ITEM_COLUMN = 2                           # Column index for Item Name (A=1, B=2, etc.)
IMAGE_COLUMN = 16                          # Column index for Image (H=8 based on visual)
HEADER_ROW = 1                            # Row number containing headers (usually 1)
# --- End Configuration ---

def sanitize_filename(name):
    """Removes invalid characters but leaves spaces intact for use as a filename."""
    if not isinstance(name, str):
        name = str(name) # Ensure it's a string
    name = name.strip()
    # Remove characters invalid for filenames in most OS
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    return name

def extract_images():
    print(f"Starting image extraction from: {EXCEL_FILE_PATH}")
    print(f"Outputting images to: {OUTPUT_IMAGE_DIR}")

    # Ensure output directory exists and clear it before saving images
    os.makedirs(OUTPUT_IMAGE_DIR, exist_ok=True)
    for filename in os.listdir(OUTPUT_IMAGE_DIR):  # New: Clear output folder
        file_path = os.path.join(OUTPUT_IMAGE_DIR, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                import shutil
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        # Changed: load the sheet by its name "Values"
        sheet = workbook["Values"]
        print(f"Successfully loaded workbook. Processing sheet: {sheet.title}")
        # Debug: Print header row and selected columns for Item and Image
        header_values = [cell.value for cell in sheet[HEADER_ROW]]
        item_header = sheet.cell(row=HEADER_ROW, column=ITEM_COLUMN).value
        image_header = sheet.cell(row=HEADER_ROW, column=IMAGE_COLUMN).value
        print("Header row values:", header_values)
        print(f"Item Column ({ITEM_COLUMN}) header:", item_header)
        print(f"Image Column ({IMAGE_COLUMN}) header:", image_header)
    except FileNotFoundError:
        print(f"Error: Excel file not found at {EXCEL_FILE_PATH}")
        return
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    # --- Image Extraction Logic ---
    # openpyxl stores images separately and anchors them to cells.
    # We need to map images back to item names via their anchor cell's row.

    images_by_row = {}
    print(f"Found {len(sheet._images)} images in the sheet.")

    for img in sheet._images:
        # Determine the row the image is anchored to
        # Anchors can be complex (TwoCellAnchor), but often _from gives the top-left cell
        row = img.anchor._from.row + 1 # openpyxl anchor rows are 0-based, sheet cells are 1-based

        # Guess image format (openpyxl doesn't always store original filename/type reliably)
        # Pillow can help identify format from bytes
        img_format = 'png' # Default to png
        try:
            img_bytes = img.ref # This gets the image data bytes
            pil_img = Image.open(io.BytesIO(img_bytes))
            detected_format = pil_img.format.lower()
            if detected_format in ['jpeg', 'jpg', 'png', 'gif', 'bmp', 'tiff']:
                 # Use standard extension, prefer jpg over jpeg
                 img_format = 'jpg' if detected_format == 'jpeg' else detected_format
            print(f"  Image anchored near row {row}, detected format: {detected_format} (using .{img_format})")
        except Exception as e:
            print(f"  Warning: Could not reliably determine format for image near row {row}. Defaulting to .png. Error: {e}")


        # Store image data and format, keyed by row
        # If multiple images are anchored to the same row (unlikely for your case?), this overwrites.
        # We assume one primary item image per row anchored near the image column.
        images_by_row[row] = {'data': img.ref, 'format': img_format}


    # --- Process Rows and Save Images ---
    saved_count = 0
    for row_idx in range(HEADER_ROW + 1, 146):  # Changed: set max row to 145
        item_name_cell = sheet.cell(row=row_idx, column=ITEM_COLUMN)
        item_name = item_name_cell.value

        if item_name and str(item_name).strip(): # Check if item name exists and is not just whitespace
            item_name = str(item_name).strip()
            print(f"Processing row {row_idx}: Item = '{item_name}'")

            # Find the image associated with this row
            if row_idx in images_by_row:
                image_info = images_by_row[row_idx]
                sanitized_name = sanitize_filename(item_name)
                output_filename = f"{sanitized_name}.{image_info['format']}"
                output_path = os.path.join(OUTPUT_IMAGE_DIR, output_filename)

                try:
                    with open(output_path, 'wb') as img_file:
                        img_file.write(image_info['data'].getvalue())  # Changed: write bytes using getvalue()
                    print(f"  -> Saved image as: {output_filename}")
                    saved_count += 1
                except IOError as e:
                    print(f"  Error: Could not save image '{output_filename}'. Reason: {e}")
                except Exception as e:
                     print(f"  Error: An unexpected error occurred saving '{output_filename}'. Reason: {e}")

            else:
                # Check adjacent rows too? Sometimes anchors are slightly off.
                # For simplicity, we first stick to exact row match.
                print(f"  Warning: No image found anchored exactly at row {row_idx} for item '{item_name}'.")

        else:
             print(f"Skipping row {row_idx}: No valid item name found in column {ITEM_COLUMN}.")


    print(f"\nExtraction complete. Saved {saved_count} images.")

if __name__ == "__main__":
    extract_images()